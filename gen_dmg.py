import openpyxl

wb = openpyxl.load_workbook("E:/Downloads/SORROW'S BF6 SCRIBBLES.xlsx", data_only=True)

skip = {'HOME','Tierlist','Attachments - Muzzles','Attachments - Barrels',
        'Attachments - Underbarrels','Attachments - Lasers & Lights',
        'Attachments - Sights','Old Versions','P18','ES 5.7','M45A1',
        'M44','GGH-22','M357 TRAIT','VZ. 61'}  # skip sidearms

shotguns = {'M87A1','M1014','18.5KS-K','DB-12'}
PELLETS = 16  # all shotguns use 16 pellets

def extract_steps(ws, pellets=None):
    """Read AA2:AB23, collapse to step-breakpoints where damage changes."""
    raw = []
    for row in ws.iter_rows(min_row=2, max_row=23, min_col=27, max_col=28, values_only=True):
        r, d = row
        if r is not None and d is not None:
            try:
                raw.append((float(r), float(d)))
            except (TypeError, ValueError):
                pass
    if not raw:
        return []

    # If pellets, divide total→per-pellet, skip zero-damage entry
    if pellets:
        raw = [(r, round(d / pellets, 2)) for r, d in raw if d > 0]

    # Collapse: keep first occurrence of each unique damage value in sequence
    steps = []
    prev_d = None
    for r, d in raw:
        if d != prev_d:
            steps.append({'r': int(r) if r == int(r) else r, 'd': d})
            prev_d = d
    return steps

for name in wb.sheetnames:
    if name in skip:
        continue
    ws = wb[name]
    pellets = PELLETS if name in shotguns else None
    steps = extract_steps(ws, pellets)
    if steps:
        js = ','.join(f"{{r:{s['r']},d:{s['d']}}}" for s in steps)
        print(f"  // {name}")
        print(f"  dmg:[{js}]")
        print()
