"""Build the BF6 attachment reference workbook from the screenshot-review JSON.

This is a human-readable view of `migration/1.3.3.0/attachment-audit/attachment-screenshot-review.json`
so the records can be browsed without reading raw JSON. It is a reference artifact: nothing
here feeds the live site, so the script does presentation only and leaves data validation to
the audit scripts under `migration/1.3.3.0/attachment-audit/`.

Replaces the older `migration/1.3.3.0/attachment-audit/build-workbook.mjs`, which required the
Codex-only `@oai/artifact-tool` package and then rewrote the resulting OOXML with regexes to
apply borders and fills. This runs anywhere Python and openpyxl are available.

    python scripts/build-attachment-workbook.py [--json PATH] [--out PATH]
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_JSON = ROOT / "migration" / "1.3.3.0" / "attachment-audit" / "attachment-screenshot-review.json"
DEFAULT_OUT = ROOT / "BF6_Attachment_Stats_Review.xlsx"

CLASS_ORDER = ["Assault Rifle", "Carbine", "SMG", "LMG", "DMR", "Sniper Rifle", "Shotgun", "Sidearm"]

# tab / title / subtitle / header / type-column colours per weapon class
PALETTE = {
    "Assault Rifle": ("1D4ED8", "172554", "DBEAFE", "1E3A8A"),
    "Carbine":       ("0F766E", "134E4A", "CCFBF1", "115E59"),
    "SMG":           ("7C3AED", "4C1D95", "EDE9FE", "5B21B6"),
    "LMG":           ("EA580C", "7C2D12", "FFEDD5", "9A3412"),
    "DMR":           ("CA8A04", "713F12", "FEF9C3", "854D0E"),
    "Sniper Rifle":  ("0284C7", "0C4A6E", "E0F2FE", "075985"),
    "Shotgun":       ("DC2626", "7F1D1D", "FEE2E2", "991B1B"),
    "Sidearm":       ("0891B2", "164E63", "CFFAFE", "155E75"),
}

BAND_FILL = "FFE2E2E2"
ROW_FILL = "FFFFFFFF"
BLOCK_RULE = "FF9CA3AF"
OVERVIEW_HEAD = "FF1E3A8A"
BUFF = "FF008000"
PENALTY = "FFFF0000"
WHITE = "FFFFFFFF"
MUTED = "FF9CA3AF"
LINK = "FF1D4ED8"

# Column A is tinted by attachment type so the blocks stay distinguishable while scrolling.
# One hue family per type, all at the same lightness so no block dominates.
TYPE_PALETTE = {
    "Muzzle":           ("FFDBEAFE", "FF1E3A8A"),  # blue
    "Barrel":           ("FFDCFCE7", "FF14532D"),  # green
    "Grip":             ("FFFEF3C7", "FF78350F"),  # amber
    "Magazine":         ("FFFCE7F3", "FF831843"),  # pink
    "Ammo":             ("FFF3E8FF", "FF581C87"),  # purple
    "Ergonomics":       ("FFE0E7FF", "FF312E81"),  # indigo
    "Laser":            ("FFCCFBF1", "FF134E4A"),  # teal
    "Light":            ("FFFEF9C3", "FF713F12"),  # yellow
    "Laser/Light":      ("FFD9F99D", "FF365314"),  # lime
    "Grip/Laser/Light": ("FFFDE68A", "FF78350F"),  # deeper amber
    "Sight":            ("FFFFE4E6", "FF881337"),  # rose
    "Range Finder":     ("FFCFFAFE", "FF164E63"),  # cyan
}
TYPE_FALLBACK = ("FFF1F5F9", "FF334155")


def type_style(attachment_type):
    return TYPE_PALETTE.get(attachment_type, TYPE_FALLBACK)


def argb(colour):
    """openpyxl writes a 6-digit colour as 00RRGGBB; Excel's own files use FFRRGGBB."""
    return colour if len(colour) == 8 else "FF" + colour

# (header, record accessor, statComparisons key, column width)
COLUMNS = [
    ("Attachment Type", lambda r: r.get("attachmentType"), None, 11.375),
    ("Attachment Name", lambda r: r.get("attachmentName"), None, 15.125),
    ("Attachment Subtype", lambda r: r.get("attachmentSubtype"), None, 11.625),
    ("Attachment Cost", lambda r: r.get("attachmentCost"), None, 10.75),
    ("Attachment Description", lambda r: r.get("attachmentDescription"), None, 37.375),
    ("Damage", lambda r: _stat(r, "damage"), "damage", 8.0),
    ("Rate of Fire (RPM)", lambda r: _stat(r, "rateOfFireRpm"), "rateOfFireRpm", 10.25),
    ("Magazine Size", lambda r: _stat(r, "magazineSize"), "magazineSize", 9.375),
    ("Hipfire", lambda r: _stat(r, "hipfire"), "hipfire", 6.875),
    ("Precision", lambda r: _stat(r, "precision"), "precision", 9.25),
    ("Control", lambda r: _stat(r, "control"), "control", 7.5),
    ("Mobility", lambda r: _stat(r, "mobility"), "mobility", 7.75),
    ("Fire Modes", lambda r: _fire_modes(r), None, 14.0),
    ("Reload Time (seconds)", lambda r: _stat(r, "reloadTimeSeconds"), "reloadTimeSeconds", 9.625),
    ("Muzzle Velocity (m/s)", lambda r: _stat(r, "muzzleVelocityMps"), "muzzleVelocityMps", 7.75),
    ("ADS Time (ms)", lambda r: _stat(r, "adsTimeMs"), "adsTimeMs", 9.5),
    ("Headshot Multiplier", lambda r: _stat(r, "headshotMultiplier"), "headshotMultiplier", 9.375),
    ("Long Range Damage", lambda r: _stat(r, "longRangeDamage"), "longRangeDamage", 8.0),
    ("3D Spot-on-fire Range (m)", lambda r: _stat(r, "spotOnFire3dM"), "spotOnFire3dM", 10.125),
    ("2D Spot-on-fire Range (m)", lambda r: _stat(r, "spotOnFire2dM"), "spotOnFire2dM", 10.125),
    ("Opponent Health Regen Delay (seconds)", lambda r: _stat(r, "opponentHealthRegenDelaySeconds"),
     "opponentHealthRegenDelaySeconds", 15.125),
    ("Collateral Multiplier", lambda r: _stat(r, "collateralMultiplier"), "collateralMultiplier", 9.375),
    ("Reload in ADS", lambda r: _reload_in_ads(r), None, 9.25),
    ("ADS Move Speed Multiplier", lambda r: _stat(r, "adsMoveSpeedMultiplier"), "adsMoveSpeedMultiplier", 10.125),
    ("Sprint Recovery (ms)", lambda r: _stat(r, "sprintRecoveryMs"), "sprintRecoveryMs", 9.375),
    ("Recoil Amount (degrees)", lambda r: _stat(r, "recoilAmountDegrees"), "recoilAmountDegrees", 9.5),
    ("Recoil Variation (degrees)", lambda r: _stat(r, "recoilVariationDegrees"), "recoilVariationDegrees", 8.88),
    ("Current Screenshot Filename", lambda r: _basename(r["source"].get("currentPath")), None, 39.375),
]

OVERVIEW_TYPE_ORDER = ["Muzzle", "Barrel", "Light", "Laser", "Laser/Light", "Grip/Laser/Light",
                       "Grip", "Magazine", "Ammo", "Ergonomics"]
SUBTYPE_KEYED_TYPES = {"Barrel", "Ammo"}


def _stat(record, key):
    stats = record.get("stats") or {}
    return stats.get(key)


def _fire_modes(record):
    modes = (record.get("stats") or {}).get("fireModes")
    if not modes:
        return None
    # seven AK-205 laser records store this as "AUTO/SINGLE" instead of an array
    if isinstance(modes, str):
        modes = [part.strip() for part in modes.split("/") if part.strip()]
    return " / ".join(modes)


def _reload_in_ads(record):
    value = (record.get("stats") or {}).get("reloadInAds")
    return None if value is None else ("Yes" if value else "No")


def _basename(path):
    return re.split(r"[\\/]", path)[-1] if path else None


def _weapon_class(record):
    parts = re.split(r"[\\/]", record["source"].get("currentPath") or "")
    if "Weapon Attachments" in parts:
        index = parts.index("Weapon Attachments")
        if index + 2 < len(parts):
            return parts[index + 1]
    return None


def load_records(json_path):
    data = json.loads(json_path.read_text(encoding="utf8"))
    records = data["records"]
    weapons = {}
    for record in records:
        # the per-weapon context capture carries no stats; it stays in Source Index only
        if record.get("attachmentType") == "Overview":
            continue
        weapons.setdefault(record["weaponName"], []).append(record)
    classes = {}
    # Rows are grouped by attachment type in the same order the Overview uses, then by capture
    # order within the type, so a weapon sheet and the Overview read top-to-bottom the same way.
    type_rank = {name: index for index, name in enumerate(OVERVIEW_TYPE_ORDER)}
    for name, group in weapons.items():
        group.sort(key=lambda r: (type_rank.get(r.get("attachmentType"), len(type_rank)),
                                  r.get("attachmentType") or "",
                                  r["source"].get("captureOrder") or 0))
        found = next((_weapon_class(r) for r in group if _weapon_class(r)), None)
        classes[name] = found or "Assault Rifle"
    order = sorted(weapons, key=lambda n: (CLASS_ORDER.index(classes[n])
                                           if classes[n] in CLASS_ORDER else len(CLASS_ORDER), n))
    return data, weapons, classes, order


def _fill(hex_colour):
    return PatternFill("solid", fgColor=argb(hex_colour))


def _option_value(record):
    """The value that identifies an option within its type: subtype for the subtype-keyed
    types, attachment name for the rest. Same rule the Overview rows are keyed on."""
    attachment_type = record.get("attachmentType")
    return (record.get("attachmentSubtype") if attachment_type in SUBTYPE_KEYED_TYPES
            else record.get("attachmentName"))


def _displayed(record, getter, cmp_key):
    """Cell value as the weapon sheets render it: an arrow-prefixed string where the game
    showed a comparison arrow, the bare value otherwise."""
    value = getter(record)
    comparison = (record.get("statComparisons") or {}).get(cmp_key) if cmp_key else None
    if comparison and value is not None:
        return ("↑" if comparison.get("direction") == "up" else "↓") + str(value)
    return value


def write_weapon_sheet(workbook, weapon, weapon_class, records):
    tab, dark, light, light_text = PALETTE.get(weapon_class, PALETTE["Assault Rifle"])
    sheet = workbook.create_sheet(weapon[:31])
    sheet.sheet_properties.tabColor = argb(tab)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "D5"

    columns = len(COLUMNS)
    title = sheet.cell(1, 1, f"{weapon} Attachment Screenshot Review")
    title.font = Font(sz=14, bold=True, color=WHITE)
    title.alignment = Alignment(vertical="center", wrap_text=False)
    subtitle = sheet.cell(2, 1, f"{weapon_class} | All values remain provisional-review-required. "
                                "Nulls carry field-specific transcription reasons; no screenshot "
                                "value is promoted to live site data.")
    subtitle.font = Font(sz=10, italic=True, color=argb(light_text))
    subtitle.alignment = Alignment(vertical="center", wrap_text=False)
    # rows 1 and 2 stay unmerged so whole columns remain selectable
    for column in range(1, columns + 1):
        sheet.cell(1, column).fill = _fill(dark)
        sheet.cell(2, column).fill = _fill(light)
    sheet.row_dimensions[1].height = 32
    sheet.row_dimensions[2].height = 24
    sheet.row_dimensions[4].height = 48

    for index, (header, _getter, _cmp, width) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(4, index, header)
        cell.fill = _fill(tab)
        cell.font = Font(sz=10, bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="bottom", horizontal="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = width

    thin_rule = Side(style="medium", color=BLOCK_RULE)
    row_for_record = {}
    for offset, record in enumerate(records):
        row = 5 + offset
        row_for_record[id(record)] = row
        banded = offset % 2 == 1
        comparisons = record.get("statComparisons") or {}
        last_of_block = (offset + 1 == len(records)
                         or records[offset + 1].get("attachmentType") != record.get("attachmentType"))
        for index, (_header, getter, cmp_key, _width) in enumerate(COLUMNS, start=1):
            value = getter(record)
            cell = sheet.cell(row, index)
            comparison = comparisons.get(cmp_key) if cmp_key else None
            if comparison and value is not None:
                arrow = "↑" if comparison.get("direction") == "up" else "↓"
                cell.value = f"{arrow}{value}"
                colour = BUFF if comparison.get("effect") == "buff" else PENALTY
                cell.font = Font(sz=10, bold=True, color=argb(colour))
                cell.alignment = Alignment(horizontal="right", vertical="top")
            else:
                cell.value = value
                cell.font = Font(sz=10)
                wrap = index in (5, 28)
                cell.alignment = Alignment(vertical="top", wrap_text=wrap)
            if index == 1:
                type_fill, type_text = type_style(record.get("attachmentType"))
                cell.fill = _fill(type_fill)
                cell.font = Font(sz=10, bold=True, color=type_text)
            else:
                cell.fill = _fill(BAND_FILL if banded else ROW_FILL)
            if last_of_block:
                cell.border = Border(bottom=thin_rule)
    return row_for_record


def read_existing_option_order(path):
    """Row order the reader is already used to, taken from the workbook being replaced.

    Options that still exist keep their established position; genuinely new ones are
    appended to their type. Missing or unreadable workbook just means first-seen order.
    """
    if not path.exists():
        return {}
    try:
        from openpyxl import load_workbook
        previous = load_workbook(path, read_only=True)
        if "Overview" not in previous.sheetnames:
            return {}
        established = {}
        for row in previous["Overview"].iter_rows(min_row=3, max_col=3, values_only=True):
            attachment_type, _field, value = (list(row) + [None, None, None])[:3]
            if attachment_type in OVERVIEW_TYPE_ORDER and value:
                established.setdefault(attachment_type, []).append(str(value))
        previous.close()
        return established
    except Exception as error:  # a corrupt or open workbook must not block a rebuild
        print(f"  note: could not read existing overview order ({error}); using first-seen order")
        return {}


def _link_label(cost):
    """HYPERLINK friendly-name argument for an overview cell: the attachment cost.

    A whole-number cost goes in bare so the cell reads as a number; anything else (a missing
    cost, or a non-numeric capture) falls back to a quoted string so the formula still parses.
    """
    if isinstance(cost, bool) or cost is None:
        return '"?"'
    if isinstance(cost, int):
        return str(cost)
    if isinstance(cost, float) and cost.is_integer():
        return str(int(cost))
    text = str(cost).strip()
    if not text:
        return '"?"'
    return '"' + text.replace('"', '""') + '"'


def write_overview(workbook, weapons, classes, order, row_lookup, established_order):
    sheet = workbook.create_sheet("Overview", 0)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "D3"

    # current options per type, in first-seen capture order
    first_seen = {attachment_type: [] for attachment_type in OVERVIEW_TYPE_ORDER}
    present = {attachment_type: set() for attachment_type in OVERVIEW_TYPE_ORDER}
    for weapon in order:
        for record in weapons[weapon]:
            attachment_type = record.get("attachmentType")
            if attachment_type not in present:
                continue
            value = _option_value(record)
            if value is None or value in present[attachment_type]:
                continue
            present[attachment_type].add(value)
            first_seen[attachment_type].append(value)

    rows = []
    for attachment_type in OVERVIEW_TYPE_ORDER:
        field = "Attachment Subtype" if attachment_type in SUBTYPE_KEYED_TYPES else "Attachment Name"
        retained = [v for v in established_order.get(attachment_type, []) if v in present[attachment_type]]
        retained_set = set(retained)
        appended = [v for v in first_seen[attachment_type] if v not in retained_set]
        for value in retained + appended:
            rows.append((attachment_type, field, value))

    # first record per weapon for each option, for the hyperlink target
    target = {}
    for weapon in order:
        for record in weapons[weapon]:
            attachment_type = record.get("attachmentType")
            if attachment_type not in OVERVIEW_TYPE_ORDER:
                continue
            value = _option_value(record)
            # the display text is the attachment cost, not "Link": an inconsistent cost between
            # weapons on the same row is then visible at a glance, which is how the sniper grip
            # variants were caught. The hyperlink target is unchanged.
            target.setdefault((weapon, attachment_type, value),
                              (row_lookup[weapon][id(record)], record.get("attachmentCost")))

    # Row 1 labels are centred with Center Across Selection rather than merged cells: the label
    # spans its group visually while every cell stays individually selectable, so whole columns
    # can still be highlighted. In OOXML that is horizontal="centerContinuous" applied to every
    # cell of the span, with the text living in the first cell only.
    header = sheet.cell(1, 1, "Attachment Options")
    header.font = Font(sz=12, bold=True, color=WHITE)
    for column in range(1, 4):
        sheet.cell(1, column).fill = _fill(OVERVIEW_HEAD)
        sheet.cell(1, column).alignment = Alignment(horizontal="centerContinuous", vertical="center")
    for index, label in enumerate(["Attachment Type", "Field", "Attachment Name / Subtype"], start=1):
        cell = sheet.cell(2, index, label)
        cell.fill = _fill(OVERVIEW_HEAD)
        cell.font = Font(sz=10, bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    previous_class = None
    for offset, weapon in enumerate(order):
        column = 4 + offset
        weapon_class = classes[weapon]
        tab = PALETTE.get(weapon_class, PALETTE["Assault Rifle"])[0]
        if weapon_class != previous_class:
            # label goes in the first cell of the group and centres across it, see above
            label = sheet.cell(1, column, weapon_class)
            label.font = Font(sz=11, bold=True, color=WHITE)
            previous_class = weapon_class
        sheet.cell(1, column).fill = _fill(tab)
        sheet.cell(1, column).alignment = Alignment(horizontal="centerContinuous", vertical="center")
        cell = sheet.cell(2, column, weapon)
        cell.fill = _fill(tab)
        cell.font = Font(sz=10, bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
        sheet.column_dimensions[get_column_letter(column)].width = 7.88

    # last column of each weapon-class group gets a vertical rule
    class_edges = {4 + offset for offset, weapon in enumerate(order)
                   if offset + 1 == len(order) or classes[order[offset + 1]] != classes[weapon]}
    rule = Side(style="medium", color=BLOCK_RULE)
    last_column = 3 + len(order)

    for offset, (attachment_type, field, value) in enumerate(rows):
        row = 3 + offset
        sheet.row_dimensions[row].height = 14.25
        banded = offset % 2 == 1
        # horizontal rule under the final row of each attachment-type block
        end_of_block = offset + 1 == len(rows) or rows[offset + 1][0] != attachment_type
        type_fill, type_text = type_style(attachment_type)

        def edges(column):
            return Border(bottom=rule if end_of_block else None,
                          right=rule if column in class_edges else None)

        for index, text in enumerate([attachment_type, field, value], start=1):
            cell = sheet.cell(row, index, text)
            cell.alignment = Alignment(vertical="center")
            if index == 1:
                cell.fill = _fill(type_fill)
                cell.font = Font(sz=10, bold=True, color=type_text)
            else:
                cell.font = Font(sz=10)
                if banded:
                    cell.fill = _fill(BAND_FILL)
            cell.border = edges(index)
        for weapon_offset, weapon in enumerate(order):
            column = 4 + weapon_offset
            cell = sheet.cell(row, column)
            hit = target.get((weapon, attachment_type, value))
            if hit:
                hit_row, cost = hit
                cell.value = f"=HYPERLINK(\"#'{weapon}'!B{hit_row}\",{_link_label(cost)})"
                cell.font = Font(sz=10, color=LINK, underline="single")
            else:
                cell.value = "—"
                cell.font = Font(sz=10, color=MUTED)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if banded:
                cell.fill = _fill(BAND_FILL)
            cell.border = edges(column)

    # carry the class rule up through the two header rows so the groups read as columns
    for header_row in (1, 2):
        for column in range(1, last_column + 1):
            if column in class_edges:
                cell = sheet.cell(header_row, column)
                cell.border = Border(right=rule)

    sheet.column_dimensions["A"].width = 21
    sheet.column_dimensions["B"].width = 15.25
    sheet.column_dimensions["C"].width = 19
    sheet.row_dimensions[1].height = 15.75
    sheet.row_dimensions[2].height = 42
    return rows


def _range_name(attachment_type):
    """Defined-name suffix for a type's option list. The only character the type names carry
    that a name cannot is the slash, so dropping it is enough to keep these unique."""
    return "opt_" + attachment_type.replace("/", "")


def write_lookup_data(workbook, weapons, order, option_rows):
    """Hidden backing sheet for the By Attachment view: one row per record keyed on
    weapon/type/option, plus the option lists the two dropdowns validate against."""
    sheet = workbook.create_sheet("Lookup Data")
    sheet.sheet_state = "hidden"

    sheet.cell(1, 1, "Key")
    for index, (header, _getter, _cmp, _width) in enumerate(COLUMNS, start=2):
        sheet.cell(1, index, header)
    row = 1
    for weapon in order:
        for record in weapons[weapon]:
            value = _option_value(record)
            if value is None:
                continue
            row += 1
            sheet.cell(row, 1, f"{weapon}|{record.get('attachmentType')}|{value}")
            for index, (_header, getter, cmp_key, _width) in enumerate(COLUMNS, start=2):
                sheet.cell(row, index, _displayed(record, getter, cmp_key))
    last_record_row = row

    # Option lists live to the right of the records, one column per type, each published as a
    # defined name so the dependent dropdown can reach it through INDIRECT.
    by_type = {}
    for attachment_type, _field, value in option_rows:
        by_type.setdefault(attachment_type, []).append(value)
    first_list_column = len(COLUMNS) + 3
    types = [t for t in OVERVIEW_TYPE_ORDER if by_type.get(t)]
    for offset, attachment_type in enumerate(types):
        column = first_list_column + offset
        letter = get_column_letter(column)
        sheet.cell(1, column, attachment_type)
        for value_offset, value in enumerate(by_type[attachment_type], start=2):
            sheet.cell(value_offset, column, value)
        workbook.defined_names.add(DefinedName(
            _range_name(attachment_type),
            attr_text=f"'Lookup Data'!${letter}$2:${letter}${len(by_type[attachment_type]) + 1}"))
    type_column = get_column_letter(first_list_column + len(types) + 1)
    sheet.cell(1, first_list_column + len(types) + 1, "Attachment Type")
    for offset, attachment_type in enumerate(types, start=2):
        sheet.cell(offset, first_list_column + len(types) + 1, attachment_type)
    workbook.defined_names.add(DefinedName(
        "opt_types", attr_text=f"'Lookup Data'!${type_column}$2:${type_column}${len(types) + 1}"))
    return types, by_type, last_record_row


def write_by_attachment(workbook, classes, order, types, by_type, record_rows):
    """One attachment across the whole roster: pick a type and an option, read a row per weapon.

    Every cell is a formula against the hidden Lookup Data sheet, so the view follows the
    dropdowns without a rebuild and picks up new weapons or options on the next one.
    """
    sheet = workbook.create_sheet("By Attachment", 1)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "B5"
    columns = len(COLUMNS) + 1
    default_type = types[0]

    for row, label, value, list_formula in (
        (1, "Attachment Type", default_type, "=opt_types"),
        (2, "Attachment Name / Subtype", by_type[default_type][0],
         '=INDIRECT("opt_"&SUBSTITUTE($B$1,"/",""))'),
    ):
        key = sheet.cell(row, 1, label)
        key.font = Font(sz=10, bold=True, color=WHITE)
        key.fill = _fill(OVERVIEW_HEAD)
        picker = sheet.cell(row, 2, value)
        picker.font = Font(sz=10, bold=True)
        picker.alignment = Alignment(horizontal="left", vertical="center")
        picker.border = Border(*(Side(style="thin", color=OVERVIEW_HEAD),) * 4)
        validation = DataValidation(type="list", formula1=list_formula, allow_blank=False)
        sheet.add_data_validation(validation)
        validation.add(picker)

    hint = sheet.cell(3, 1, "Pick a type and an option above. Every weapon gets a row; blank "
                            "means that weapon does not offer the option.")
    hint.font = Font(sz=9, italic=True, color=argb(MUTED))

    header = sheet.cell(4, 1, "Weapon")
    header.fill = _fill(OVERVIEW_HEAD)
    header.font = Font(sz=10, bold=True, color=WHITE)
    header.alignment = Alignment(vertical="bottom", horizontal="center", wrap_text=True)
    for index, (text, _getter, _cmp, width) in enumerate(COLUMNS, start=2):
        cell = sheet.cell(4, index, text)
        cell.fill = _fill(OVERVIEW_HEAD)
        cell.font = Font(sz=10, bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="bottom", horizontal="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.column_dimensions["A"].width = 21
    sheet.row_dimensions[4].height = 48

    keys = f"'Lookup Data'!$A$2:$A${record_rows}"
    rule = Side(style="medium", color=BLOCK_RULE)
    row = 4
    previous_class = None
    for weapon in order:
        weapon_class = classes[weapon]
        if weapon_class != previous_class:
            row += 1
            tab = PALETTE.get(weapon_class, PALETTE["Assault Rifle"])[0]
            band = sheet.cell(row, 1, weapon_class)
            band.font = Font(sz=11, bold=True, color=WHITE)
            for column in range(1, columns + 1):
                sheet.cell(row, column).fill = _fill(tab)
                sheet.cell(row, column).alignment = Alignment(horizontal="centerContinuous",
                                                              vertical="center")
                sheet.cell(row, column).border = Border(bottom=rule)
            previous_class = weapon_class
        row += 1
        light, light_text = PALETTE.get(weapon_class, PALETTE["Assault Rifle"])[2:4]
        name = sheet.cell(row, 1, weapon)
        name.fill = _fill(light)
        name.font = Font(sz=10, bold=True, color=argb(light_text))
        name.alignment = Alignment(vertical="top")
        for index in range(2, columns + 1):
            source = f"'Lookup Data'!${get_column_letter(index)}$2:${get_column_letter(index)}${record_rows}"
            cell = sheet.cell(row, index,
                              f'=IFERROR(INDEX({source},MATCH($A{row}&"|"&$B$1&"|"&$B$2,{keys},0)),"")')
            cell.font = Font(sz=10)
            cell.alignment = Alignment(vertical="top", wrap_text=index in (6, 29))

    last_row = row
    # The arrow prefix is what the weapon sheets carry; colour it the same way here.
    value_range = f"B5:{get_column_letter(columns)}{last_row}"
    for arrow, colour in (("↑", BUFF), ("↓", PENALTY)):
        sheet.conditional_formatting.add(value_range, FormulaRule(
            formula=[f'AND(ISTEXT(B5),LEFT(B5,1)="{arrow}")'],
            font=Font(sz=10, bold=True, color=argb(colour)), stopIfTrue=False))
    return last_row - 4


def write_source_index(workbook, records):
    sheet = workbook.create_sheet("Source Index")
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    headers = ["Weapon", "Attachment Type", "Attachment Subtype", "Attachment Cost", "Attachment Name",
               "Extraction Status", "Review Status", "Mapping Review Status", "Current Screenshot Path",
               "Original Screenshot Path", "Capture Timestamp", "Resolution"]
    widths = [16, 16, 16, 12, 22, 22, 26, 20, 60, 60, 16, 12]
    for index, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = sheet.cell(1, index, header)
        cell.fill = _fill(OVERVIEW_HEAD)
        cell.font = Font(sz=10, bold=True, color=WHITE)
        sheet.column_dimensions[get_column_letter(index)].width = width
    ordered = sorted(records, key=lambda r: (r["weaponName"], r["source"].get("captureOrder") or 0))
    for offset, record in enumerate(ordered):
        row = 2 + offset
        source = record["source"]
        values = [record["weaponName"], record.get("attachmentType"), record.get("attachmentSubtype"),
                  record.get("attachmentCost"), record.get("attachmentName"), record.get("extractionStatus"),
                  record.get("reviewStatus"), record.get("mappingReviewStatus"),
                  source.get("currentPath"), source.get("originalPath"),
                  source.get("captureTimestamp"), source.get("resolution")]
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row, index, value)
            cell.font = Font(sz=10)
        # Column 9 opens the capture itself: the filename reads better than the absolute path,
        # and the path is still there as the link target.
        current = source.get("currentPath")
        if current:
            cell = sheet.cell(row, 9, os.path.basename(current))
            # Leave the drive colon and separators literal; Excel wants file:///C:/... with only
            # spaces and other unsafe characters percent-encoded.
            cell.hyperlink = f"file:///{quote(current.replace(chr(92), '/'), safe='/:')}"
            cell.font = Font(sz=10, color=LINK, underline="single")
    sheet.column_dimensions["I"].width = 46
    return len(ordered)


def write_read_me(workbook, data, weapon_count, out_path, json_path):
    sheet = workbook.create_sheet("Read Me")
    sheet.sheet_view.showGridLines = False
    title = sheet.cell(1, 1, "BF6 Attachment Workbook - Read Me")
    title.font = Font(sz=14, bold=True, color=WHITE)
    title.fill = _fill(OVERVIEW_HEAD)
    sheet.cell(1, 2).fill = _fill(OVERVIEW_HEAD)
    entries = [
        ("Purpose", "Human-readable view of the screenshot audit. Reference only - no value here "
                    "feeds the live site."),
        ("Source JSON", str(json_path.relative_to(ROOT)) if json_path.is_relative_to(ROOT) else str(json_path)),
        ("JSON schema", str(data.get("schemaVersion"))),
        ("Source generatedAt", str(data.get("generatedAt"))),
        ("Records", str(data.get("recordCount"))),
        ("Attachment detail records", str(data.get("attachmentDetailCount"))),
        ("Weapons in workbook", str(weapon_count)),
        ("Status", str(data.get("status"))),
        ("Built by", "scripts/build-attachment-workbook.py"),
        ("Built at", time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())),
        ("", ""),
        ("Reading the sheets", "Each weapon sheet lists one row per attachment-detail screenshot in "
                               "capture order. Panes are frozen at D5."),
        ("Change indicators", "A stat that the game showed with a comparison arrow is rendered as "
                              "an arrow plus the displayed value: bold green for a buff, bold red "
                              "for a penalty. The arrow is what the screen showed; it does not "
                              "always mean the displayed number differs from the baseline."),
        ("Blank cells", "A blank stat means the value was not readable in that screenshot. The "
                        "reason is recorded per field in the source JSON, not here."),
        ("Overview sheet", "Attachment option matrix. Each cell links to that weapon's row."),
        ("By Attachment sheet", "One attachment across the whole roster. Pick an attachment type "
                                "in B1 and an option in B2; each weapon's row fills in from the "
                                "hidden Lookup Data sheet. A blank row means that weapon does not "
                                "offer the option."),
    ]
    for offset, (label, value) in enumerate(entries):
        row = 3 + offset
        key = sheet.cell(row, 1, label or None)
        key.font = Font(sz=10, bold=True)
        key.alignment = Alignment(vertical="top")
        body = sheet.cell(row, 2, value or None)
        body.font = Font(sz=10)
        body.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 96


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--json", type=Path, default=DEFAULT_JSON)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()

    started = time.time()
    data, weapons, classes, order = load_records(args.json)
    established_order = read_existing_option_order(args.out)

    workbook = Workbook()
    workbook.remove(workbook.active)

    row_lookup = {}
    for weapon in order:
        row_lookup[weapon] = write_weapon_sheet(workbook, weapon, classes[weapon], weapons[weapon])
    option_rows = write_overview(workbook, weapons, classes, order, row_lookup, established_order)
    types, by_type, record_rows = write_lookup_data(workbook, weapons, order, option_rows)
    lookup_rows = write_by_attachment(workbook, classes, order, types, by_type, record_rows)
    index_rows = write_source_index(workbook, data["records"])
    write_read_me(workbook, data, len(order), args.out, args.json)

    args.out.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.out)

    print(f"wrote {args.out}")
    print(f"  weapons {len(order)} | sheets {len(workbook.sheetnames)} | "
          f"records {len(data['records'])} | overview options {len(option_rows)} | "
          f"by-attachment rows {lookup_rows} | index rows {index_rows}")
    print(f"  {time.time() - started:.1f}s")


if __name__ == "__main__":
    sys.exit(main())
