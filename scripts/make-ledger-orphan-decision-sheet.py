"""Build the decision workbook for the override-ledger entries that are still orphaned.

After the two 2026-08-01 path-reconciliation rounds, 286 entries in
`migration/1.3.3.0/attachment-audit/manual-review-overrides.json` still point at screenshot filenames that
no longer exist. Each one needs a human call rather than a key repair, so this renders them with
the before/after each choice implies: what the record holds today, what the live entry already
sets, and what the orphan would set if it were revived.

Reporting only — it reads the audit JSON and writes an xlsx, and changes no data.

    python scripts/make-ledger-orphan-decision-sheet.py [--out PATH]
"""

from __future__ import annotations

import argparse
import json
import os
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
REVIEW = ROOT / "migration" / "1.3.3.0" / "attachment-audit" / "attachment-screenshot-review.json"
LEDGER = ROOT / "migration" / "1.3.3.0" / "attachment-audit" / "manual-review-overrides.json"
DEFAULT_OUT = ROOT / "BF6_Ledger_Orphan_Decisions.xlsx"

HEAD = "FF1E3A8A"
WHITE = "FFFFFFFF"
BAND = "FFF3F4F6"
MUTED = "FF6B7280"

# one colour per decision class, mirroring the review workbook's per-class tinting
CLASS_STYLE = {
    "redundant-duplicate":   ("FF15803D", "Values already true of the record; carries no information"),
    "conflicts-with-live":   ("FFB91C1C", "Disagrees with the live entry; needs per-field adjudication"),
    "superseded":            ("FFC2410C", "Would regress the record; superseded by the 2026-07-28 repairs"),
    "kord-consolidation":    ("FF7C3AED", "Pre-consolidation KORD capture; would undo Light+Laser -> Laser/Light"),
    "no-target":             ("FF4B5563", "Capture no longer exists; nothing to re-key onto"),
}

RECOMMENDATION = {
    "redundant-duplicate": "Delete, or keep for provenance — no data effect either way",
    "conflicts-with-live": "Adjudicate per field against the screenshot",
    "superseded":          "Leave dead — reviving regresses the record",
    "kord-consolidation":  "Leave dead, or strip updates.attachmentType then re-key",
    "no-target":           "Delete — no record to attach to",
}

strip_prefix = lambda name: re.sub(r"^\d+_", "", name or "").lower()


def field_value(record, field):
    """A record field lives at the top level or inside stats."""
    if field in record:
        return record[field]
    return (record.get("stats") or {}).get(field)


def flatten(updates):
    """updates may nest a partial stats object; yield (field, value) pairs either way."""
    for field, value in (updates or {}).items():
        if field == "stats" and isinstance(value, dict):
            for stat_field, stat_value in value.items():
                yield stat_field, stat_value
        else:
            yield field, value


def classify(entry, record, occupant):
    if record is None:
        return "no-target"
    if entry.get("weaponName") == "KORD 6P67" and entry.get("attachmentType") in ("Light", "Laser"):
        return "kord-consolidation"
    changes = [f for f, v in flatten(entry.get("updates")) if field_value(record, f) != v]
    if occupant is None:
        return "superseded" if changes else "rekeyable"
    return "conflicts-with-live" if changes else "redundant-duplicate"


def collect():
    review = json.loads(REVIEW.read_text(encoding="utf-8"))
    ledger = json.loads(LEDGER.read_text(encoding="utf-8"))

    by_suffix = defaultdict(list)
    for record in review["records"]:
        by_suffix[strip_prefix(os.path.basename(record["source"]["currentPath"]))].append(record)
    # KORD's separate Light and Laser captures were consolidated into one Laser/Light selector,
    # so their filenames no longer match; they resolve on weapon + attachment name instead.
    by_name = {(r["weaponName"], r.get("attachmentName")): r
               for r in review["records"] if r.get("attachmentType") == "Laser/Light"}
    occupied = {e["sourcePath"]: e for e in ledger["overrides"]}

    rows, conflicts = [], []
    for entry in ledger["overrides"]:
        if os.path.isfile(entry["sourcePath"]):
            continue
        matches = by_suffix.get(strip_prefix(entry["sourceFilename"]), [])
        record = matches[0] if len(matches) == 1 else None
        if record is None and entry.get("weaponName") == "KORD 6P67":
            record = by_name.get((entry["weaponName"], entry.get("attachmentName")))
        occupant = occupied.get(record["source"]["currentPath"]) if record else None
        if occupant is entry:
            occupant = None
        kind = classify(entry, record, occupant)

        differing = []
        if record is not None:
            live = dict(flatten(occupant.get("updates"))) if occupant else {}
            for field, value in flatten(entry.get("updates")):
                current = field_value(record, field)
                if current == value:
                    continue
                differing.append(field)
                conflicts.append({
                    "class": kind,
                    "weapon": entry.get("weaponName"),
                    "attachment": entry.get("attachmentName"),
                    "type": record.get("attachmentType"),
                    "field": field,
                    "record_today": current,
                    "live_sets": live.get(field, "—" if not occupant else "(not set)"),
                    "orphan_would_set": value,
                    "stale_file": entry["sourceFilename"],
                    "target_file": os.path.basename(record["source"]["currentPath"]),
                })

        rows.append({
            "class": kind,
            "weapon": entry.get("weaponName"),
            "type": entry.get("attachmentType") or (record.get("attachmentType") if record else None),
            "attachment": entry.get("attachmentName"),
            "stale_file": entry["sourceFilename"],
            "target_file": os.path.basename(record["source"]["currentPath"]) if record else "—",
            "target_occupied": "yes" if occupant else ("no" if record else "—"),
            "update_fields": len(list(flatten(entry.get("updates")))),
            "conflicting": len(differing),
            "conflicting_names": ", ".join(differing[:6]) + ("…" if len(differing) > 6 else ""),
            "recommendation": RECOMMENDATION.get(kind, "Re-key"),
        })
    return rows, conflicts


def _fill(colour):
    return PatternFill("solid", fgColor=colour)


def header_row(sheet, headers, widths):
    for index, (text, width) in enumerate(zip(headers, widths), start=1):
        cell = sheet.cell(1, index, text)
        cell.fill = _fill(HEAD)
        cell.font = Font(sz=10, bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[1].height = 30
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def write_summary(workbook, rows):
    sheet = workbook.create_sheet("Summary")
    sheet.sheet_view.showGridLines = False
    title = sheet.cell(1, 1, "Orphaned override-ledger entries — decisions needed")
    title.font = Font(sz=13, bold=True)
    sheet.cell(2, 1, f"{len(rows)} entries still point at screenshot filenames that no longer exist. "
                     "Each needs a judgment call, not a key repair.").font = Font(sz=10, color=MUTED)

    headers = ["Class", "Count", "What it means", "Recommendation"]
    for index, (text, width) in enumerate(zip(headers, [22, 8, 62, 52]), start=1):
        cell = sheet.cell(4, index, text)
        cell.fill = _fill(HEAD)
        cell.font = Font(sz=10, bold=True, color=WHITE)
        sheet.column_dimensions[get_column_letter(index)].width = width

    counts = defaultdict(int)
    for row in rows:
        counts[row["class"]] += 1
    for offset, kind in enumerate(["redundant-duplicate", "conflicts-with-live", "superseded",
                                   "kord-consolidation", "no-target"]):
        row_index = 5 + offset
        colour, meaning = CLASS_STYLE[kind]
        cell = sheet.cell(row_index, 1, kind)
        cell.fill = _fill(colour)
        cell.font = Font(sz=10, bold=True, color=WHITE)
        sheet.cell(row_index, 2, counts.get(kind, 0)).font = Font(sz=10, bold=True)
        sheet.cell(row_index, 3, meaning).font = Font(sz=10)
        sheet.cell(row_index, 4, RECOMMENDATION[kind]).font = Font(sz=10)
        for column in range(1, 5):
            sheet.cell(row_index, column).alignment = Alignment(vertical="center", wrap_text=True)
        sheet.row_dimensions[row_index].height = 28
    total = sheet.cell(5 + len(CLASS_STYLE), 1, "Total")
    total.font = Font(sz=10, bold=True)
    sheet.cell(5 + len(CLASS_STYLE), 2, len(rows)).font = Font(sz=10, bold=True)

    note = ("The hard case is the 12 SL9 Laser/Light entries inside conflicts-with-live: the live "
            "entry is right on spotOnFire2dM (150, matching every screenshot) but looks wrong on "
            "collateralMultiplier (0). Neither side is uniformly correct.")
    cell = sheet.cell(12, 1, note)
    cell.font = Font(sz=10, italic=True, color=MUTED)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells(start_row=12, start_column=1, end_row=14, end_column=4)


def write_entries(workbook, rows):
    sheet = workbook.create_sheet("Entries")
    sheet.sheet_view.showGridLines = False
    headers = ["Class", "Weapon", "Type", "Attachment", "Stale filename (dead key)",
               "Would re-key onto", "Target occupied", "Update fields", "Conflicting",
               "Conflicting fields", "Recommendation", "Your decision"]
    header_row(sheet, headers, [21, 13, 13, 24, 46, 46, 11, 9, 10, 34, 44, 18])
    rule = Side(style="thin", color="FFD1D5DB")
    for offset, row in enumerate(sorted(rows, key=lambda r: (r["class"], r["weapon"] or "", r["attachment"] or ""))):
        index = 2 + offset
        values = [row["class"], row["weapon"], row["type"], row["attachment"], row["stale_file"],
                  row["target_file"], row["target_occupied"], row["update_fields"],
                  row["conflicting"], row["conflicting_names"], row["recommendation"], ""]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(index, column, value)
            cell.font = Font(sz=10)
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if column in (7, 8, 9) else "left")
            cell.border = Border(bottom=rule)
            if offset % 2 == 1:
                cell.fill = _fill(BAND)
        tint = sheet.cell(index, 1)
        tint.fill = _fill(CLASS_STYLE[row["class"]][0])
        tint.font = Font(sz=10, bold=True, color=WHITE)
        if row["conflicting"]:
            sheet.cell(index, 9).font = Font(sz=10, bold=True, color="FFB91C1C")


def write_conflicts(workbook, conflicts):
    sheet = workbook.create_sheet("Field conflicts")
    sheet.sheet_view.showGridLines = False
    headers = ["Class", "Weapon", "Type", "Attachment", "Field", "Record holds today",
               "Live entry sets", "Orphan would set", "Stale filename", "Target filename",
               "Your decision"]
    header_row(sheet, headers, [21, 13, 14, 26, 26, 20, 18, 20, 44, 44, 18])
    rule = Side(style="thin", color="FFD1D5DB")
    order = sorted(conflicts, key=lambda c: (c["class"], c["weapon"] or "", c["attachment"] or "", c["field"]))
    for offset, item in enumerate(order):
        index = 2 + offset
        values = [item["class"], item["weapon"], item["type"], item["attachment"], item["field"],
                  item["record_today"], item["live_sets"], item["orphan_would_set"],
                  item["stale_file"], item["target_file"], ""]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(index, column, value if not isinstance(value, (dict, list)) else json.dumps(value))
            cell.font = Font(sz=10)
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if column in (6, 7, 8) else "left")
            cell.border = Border(bottom=rule)
            if offset % 2 == 1:
                cell.fill = _fill(BAND)
        tint = sheet.cell(index, 1)
        tint.fill = _fill(CLASS_STYLE[item["class"]][0])
        tint.font = Font(sz=10, bold=True, color=WHITE)
        # the value that would change is what the reader is deciding about
        sheet.cell(index, 8).font = Font(sz=10, bold=True, color="FFB91C1C")
        sheet.cell(index, 6).font = Font(sz=10, bold=True, color="FF15803D")


CONFIDENCE_STYLE = {
    "confirmed": ("FF15803D", "Signature is unambiguous or screenshot-backed"),
    "needs check": ("FFB45309", "Could be a per-weapon exception rather than an error"),
}


SUSPECTED_SPECIAL_COLLATERAL = set()

# (weapon, field, attachment) the user has checked against the game and confirmed the record is
# wrong. These carry no automatic signature, so without this they would sit in the review pile.
USER_CONFIRMED = {
    ("P18", "collateralMultiplier", "Frangible"),
    ("M44", "collateralMultiplier", "Frangible"),
    ("VZ. 61", "collateralMultiplier", "Frangible"),
    ("M433", "collateralMultiplier", "Frangible"),
    ("M433", "collateralMultiplier", "Hollow Point"),
    ("LMR27", "headshotMultiplier", "Hollow Point"),
    ("M433", "headshotMultiplier", "Hollow Point"),
    # Hollow Point follows Frangible on these three; the M45A1 keeps 0.57 on both and is
    # handled as a rule exception instead.
    ("P18", "collateralMultiplier", "Hollow Point"),
    ("P18", "collateralMultiplier", "Subsonic"),
    ("P18", "collateralMultiplier", "Subsonic HP"),
    ("M44", "collateralMultiplier", "Hollow Point"),
    ("VZ. 61", "collateralMultiplier", "Hollow Point"),
    ("VZ. 61", "headshotMultiplier", "Synthetic Tip"),
}


def classify_violation(item):
    """Separate the violations with a known signature from the ones needing a screenshot."""
    weapon, field, holds = item["weapon"], item["field"], item["recordHolds"]
    if (weapon, field, item["attachmentName"]) in USER_CONFIRMED:
        return "confirmed", "User checked this against the game and confirmed the record is wrong"
    if field == "headshotMultiplier" and weapon in ("EF88", "BROD 3"):
        return "confirmed", "EA-acknowledged stat-screen bug; in-game value is the class value"
    if field == "collateralMultiplier" and holds == 0:
        return "confirmed", "0 is not a valid multiplier"
    if field == "collateralMultiplier" and weapon in SUSPECTED_SPECIAL_COLLATERAL:
        return "needs check", "Reads 0.67 on Standard like PW7A2 — likely its own scale, so the target value is unknown"
    if holds == 1 and item["ammoRuleKey"] in ("Synthetic", "Penetration"):
        return "confirmed", "Reads exactly 1 — the arrow-glyph misparse (section 18.4)"
    if field == "opponentHealthRegenDelaySeconds":
        return "confirmed", "Only the ammo slot drives regen; non-ammo captures must read 5s"
    return "needs check", "No known signature — may be a per-weapon exception"


def load_violations():
    path = ROOT / "migration" / "1.3.3.0" / "attachment-audit" / "ammo-rule-violations-20260801.json"
    if not path.exists():
        return []
    items = json.loads(path.read_text(encoding="utf-8"))
    for item in items:
        item["confidence"], item["why"] = classify_violation(item)
    return items


def write_stat_rules(workbook, items):
    """One row per (weapon, field, value) group — the unit a single decision covers."""
    sheet = workbook.create_sheet("Stat rules")
    sheet.sheet_view.showGridLines = False
    headers = ["Confidence", "Class", "Weapon", "Field", "Applies to", "Record holds",
               "Rule expects", "Records", "Why", "Your decision"]
    header_row(sheet, headers, [14, 14, 13, 30, 40, 13, 13, 9, 54, 18])
    grouped = defaultdict(list)
    for item in items:
        grouped[(item["confidence"], item["class"], item["weapon"], item["field"],
                 item["recordHolds"], item["ruleExpects"], item["why"])].append(item)
    rule = Side(style="thin", color="FFD1D5DB")
    order = sorted(grouped.items(), key=lambda kv: (kv[0][0] != "confirmed", -len(kv[1]), kv[0][2]))
    for offset, (key, members) in enumerate(order):
        confidence, cls, weapon, field, holds, expects, why = key
        index = 2 + offset
        scope = sorted({str(m["attachmentName"]) for m in members})
        applies = f"all {len(members)} attachments" if len(members) > 8 else ", ".join(scope)
        values = [confidence, cls, weapon, field, applies, holds, expects, len(members), why, ""]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(index, column, value)
            cell.font = Font(sz=10)
            cell.alignment = Alignment(vertical="center", wrap_text=column == 9,
                                       horizontal="center" if column in (6, 7, 8) else "left")
            cell.border = Border(bottom=rule)
            if offset % 2 == 1:
                cell.fill = _fill(BAND)
        tint = sheet.cell(index, 1)
        tint.fill = _fill(CONFIDENCE_STYLE[confidence][0])
        tint.font = Font(sz=10, bold=True, color=WHITE)
        sheet.cell(index, 6).font = Font(sz=10, bold=True, color="FFB91C1C")
        sheet.cell(index, 7).font = Font(sz=10, bold=True, color="FF15803D")


def write_stat_detail(workbook, items):
    sheet = workbook.create_sheet("Stat rule detail")
    sheet.sheet_view.showGridLines = False
    headers = ["Confidence", "Class", "Weapon", "Attachment type", "Attachment", "Ammo row",
               "Field", "Record holds", "Rule expects", "Screenshot"]
    header_row(sheet, headers, [14, 14, 13, 16, 26, 18, 30, 13, 13, 70])
    rule = Side(style="thin", color="FFD1D5DB")
    order = sorted(items, key=lambda i: (i["confidence"] != "confirmed", i["weapon"], i["field"],
                                         str(i["attachmentName"])))
    for offset, item in enumerate(order):
        index = 2 + offset
        values = [item["confidence"], item["class"], item["weapon"], item["attachmentType"],
                  item["attachmentName"], item["ammoRuleKey"], item["field"],
                  item["recordHolds"], item["ruleExpects"], os.path.basename(item["screenshot"])]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(index, column, value)
            cell.font = Font(sz=10)
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if column in (8, 9) else "left")
            cell.border = Border(bottom=rule)
            if offset % 2 == 1:
                cell.fill = _fill(BAND)
        tint = sheet.cell(index, 1)
        tint.fill = _fill(CONFIDENCE_STYLE[item["confidence"]][0])
        tint.font = Font(sz=10, bold=True, color=WHITE)
        sheet.cell(index, 8).font = Font(sz=10, bold=True, color="FFB91C1C")
        sheet.cell(index, 9).font = Font(sz=10, bold=True, color="FF15803D")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", default=str(DEFAULT_OUT))
    args = parser.parse_args()

    rows, conflicts = collect()
    violations = load_violations()
    workbook = Workbook()
    workbook.remove(workbook.active)
    write_summary(workbook, rows)
    write_entries(workbook, rows)
    write_conflicts(workbook, conflicts)
    if violations:
        write_stat_rules(workbook, violations)
        write_stat_detail(workbook, violations)
    workbook.save(args.out)
    if violations:
        counts = Counter(item["confidence"] for item in violations)
        print(f"  stat-rule violations {len(violations)} | " +
              " | ".join(f"{k} {v}" for k, v in counts.most_common()))
    print(f"wrote {args.out}")
    print(f"  entries {len(rows)} | field conflicts {len(conflicts)}")
    counts = defaultdict(int)
    for row in rows:
        counts[row["class"]] += 1
    for kind, count in sorted(counts.items(), key=lambda kv: -kv[1]):
        print(f"    {kind:<22}{count}")


if __name__ == "__main__":
    main()
