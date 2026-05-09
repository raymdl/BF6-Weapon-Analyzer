import openpyxl
import json

wb = openpyxl.load_workbook("E:/Downloads/SORROW'S BF6 SCRIBBLES.xlsx", data_only=True)
print("SHEETS:", wb.sheetnames)
print()

# For each weapon sheet, read AA2:AB23 (range/damage pairs)
results = {}
skip = {'Index','Home','Template','Notes','Changelog','Graphs'}
for name in wb.sheetnames:
    if name in skip:
        continue
    ws = wb[name]
    pairs = []
    for row in ws.iter_rows(min_row=2, max_row=23, min_col=27, max_col=28, values_only=True):
        r, d = row
        if r is not None and d is not None:
            try:
                pairs.append((float(r), float(d)))
            except (TypeError, ValueError):
                pass
    if pairs:
        results[name] = pairs

for wname, pairs in results.items():
    print(f"{wname}: {pairs}")
